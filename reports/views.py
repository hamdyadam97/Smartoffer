import io
from datetime import datetime, timedelta
from decimal import Decimal

from django.http import HttpResponse
from django.db.models import Sum, Count, Q
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework.permissions import IsAuthenticated

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.units import cm

from finance.models import Payment, PaymentOut, Offer
from registrations.models import Account
from students.models import Student
from courses.models import Course


class DashboardStatsView(APIView):
    """إحصائيات لوحة التحكم"""
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        today = datetime.now().date()
        month_start = today.replace(day=1)
        
        # إحصائيات الطلاب
        total_students = Student.objects.count()
        new_students_this_month = Student.objects.filter(
            created_at__date__gte=month_start
        ).count()
        
        # إحصائيات المدفوعات
        total_payments = Payment.objects.filter(
            type='ايرادات اساسية'
        ).aggregate(total=Sum('amount_number'))['total'] or Decimal('0')
        
        payments_this_month = Payment.objects.filter(
            type='ايرادات اساسية',
            date__date__gte=month_start
        ).aggregate(total=Sum('amount_number'))['total'] or Decimal('0')
        
        # إحصائيات عروض الأسعار
        total_offers = Offer.objects.count()
        converted_offers = Offer.objects.filter(registered=True).count()
        conversion_rate = (converted_offers / total_offers * 100) if total_offers > 0 else 0
        
        # إحصائيات التسجيلات
        total_registrations = Account.objects.count()
        registrations_this_month = Account.objects.filter(
            register_date__date__gte=month_start
        ).count()
        
        return Response({
            'students': {
                'total': total_students,
                'new_this_month': new_students_this_month
            },
            'payments': {
                'total': float(total_payments),
                'this_month': float(payments_this_month)
            },
            'offers': {
                'total': total_offers,
                'converted': converted_offers,
                'conversion_rate': round(conversion_rate, 2)
            },
            'registrations': {
                'total': total_registrations,
                'this_month': registrations_this_month
            }
        })


class FinancialReportView(APIView):
    """التقرير المالي"""
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        
        payments = Payment.objects.filter(type='ايرادات اساسية')
        payments_out = PaymentOut.objects.all()
        
        if start_date:
            payments = payments.filter(date__date__gte=start_date)
            payments_out = payments_out.filter(date__date__gte=start_date)
        if end_date:
            payments = payments.filter(date__date__lte=end_date)
            payments_out = payments_out.filter(date__date__lte=end_date)
        
        total_income = payments.aggregate(total=Sum('amount_number'))['total'] or Decimal('0')
        total_expense = payments_out.aggregate(total=Sum('amount_number'))['total'] or Decimal('0')
        
        return Response({
            'period': {
                'start': start_date,
                'end': end_date
            },
            'income': float(total_income),
            'expense': float(total_expense),
            'net': float(total_income - total_expense),
            'payments_count': payments.count(),
            'payments_out_count': payments_out.count()
        })


class ExportExcelView(APIView):
    """تصدير Excel"""
    permission_classes = [IsAuthenticated]
    
    def get(self, request, report_type):
        wb = Workbook()
        ws = wb.active
        
        # RTL للورقة
        ws.sheet_view.rightToLeft = True
        
        # تنسيق العناوين
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        
        if report_type == 'students':
            ws.title = "الطلاب"
            headers = ['الاسم', 'رقم الهوية', 'المحمول', 'البريد الإلكتروني', 'تاريخ الإضافة']
            ws.append(headers)
            
            for student in Student.objects.select_related('contact').all():
                ws.append([
                    student.get_full_name(),
                    student.contact.identity_number,
                    student.contact.mobile,
                    student.contact.identity_location or '',
                    student.created_at.strftime('%Y-%m-%d')
                ])
        
        elif report_type == 'payments':
            ws.title = "سندات القبض"
            headers = ['الكود', 'التاريخ', 'المبلغ', 'طريقة الدفع', 'الطالب', 'تاريخ الإضافة']
            ws.append(headers)
            
            for payment in Payment.objects.select_related('account', 'account__student', 'account__student__contact').all():
                ws.append([
                    payment.code,
                    payment.date.strftime('%Y-%m-%d'),
                    float(payment.amount_number),
                    payment.get_payment_method_display(),
                    payment.account.student.get_full_name(),
                    payment.created_at.strftime('%Y-%m-%d')
                ])
        
        elif report_type == 'offers':
            ws.title = "عروض الأسعار"
            headers = ['الكود', 'اسم العميل', 'الجوال', 'المبلغ', 'الخصم', 'الصافي', 'تم التسجيل']
            ws.append(headers)
            
            for offer in Offer.objects.all():
                ws.append([
                    offer.code,
                    offer.customer_name,
                    offer.customer_mobile,
                    float(offer.master_price),
                    float(offer.master_discount_amount),
                    offer.get_net(),
                    'نعم' if offer.registered else 'لا'
                ])
        
        # تنسيق الرأس
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # ضبط عرض الأعمدة
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # إنشاء الاستجابة
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{report_type}_{datetime.now().strftime("%Y%m%d")}.xlsx"'
        
        wb.save(response)
        return response


class ExportPDFView(APIView):
    """تصدير PDF"""
    permission_classes = [IsAuthenticated]
    
    def get(self, request, report_type):
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{report_type}_{datetime.now().strftime("%Y%m%d")}.pdf"'
        
        doc = SimpleDocTemplate(
            response,
            pagesize=A4,
            rightMargin=2*cm,
            leftMargin=2*cm,
            topMargin=2*cm,
            bottomMargin=2*cm
        )
        
        elements = []
        styles = getSampleStyleSheet()
        
        # عنوان التقرير
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontName='Helvetica-Bold',
            fontSize=18,
            alignment=1,  # Center
            spaceAfter=30
        )
        
        if report_type == 'students':
            elements.append(Paragraph('تقرير الطلاب', title_style))
            elements.append(Spacer(1, 20))
            
            data = [['الاسم', 'رقم الهوية', 'المحمول', 'تاريخ الإضافة']]
            for student in Student.objects.select_related('contact').all()[:100]:
                data.append([
                    student.get_full_name(),
                    student.contact.identity_number or '-',
                    student.contact.mobile or '-',
                    student.created_at.strftime('%Y-%m-%d')
                ])
        
        elif report_type == 'payments':
            elements.append(Paragraph('تقرير سندات القبض', title_style))
            elements.append(Spacer(1, 20))
            
            data = [['الكود', 'التاريخ', 'المبلغ', 'الطالب']]
            for payment in Payment.objects.select_related('account', 'account__student', 'account__student__contact').all()[:100]:
                data.append([
                    str(payment.code),
                    payment.date.strftime('%Y-%m-%d'),
                    f"{float(payment.amount_number):,.2f}",
                    payment.account.student.get_full_name()
                ])
        
        elif report_type == 'offers':
            elements.append(Paragraph('تقرير عروض الأسعار', title_style))
            elements.append(Spacer(1, 20))
            
            data = [['الكود', 'العميل', 'المبلغ', 'الصافي', 'الحالة']]
            for offer in Offer.objects.all()[:100]:
                data.append([
                    str(offer.code),
                    offer.customer_name,
                    f"{float(offer.master_price):,.2f}",
                    f"{offer.get_net():,.2f}",
                    'مسجل' if offer.registered else 'غير مسجل'
                ])
        
        # إنشاء الجدول
        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        
        elements.append(table)
        
        # تاريخ التقرير
        elements.append(Spacer(1, 30))
        date_style = ParagraphStyle(
            'DateStyle',
            parent=styles['Normal'],
            fontSize=10,
            alignment=1
        )
        elements.append(Paragraph(f'تم إنشاء التقرير في: {datetime.now().strftime("%Y-%m-%d %H:%M")}', date_style))
        
        doc.build(elements)
        return response
